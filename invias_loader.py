"""
INVIAS APU Bulk Loader v2
Reads all 140 INVIAS 2026-1 Excel files and inserts item+insumo data
directly into the `apus` DB table.

Usage:
    python invias_loader.py [--dry-run] [--limit N]
"""

import argparse
import logging
import os
import re
import time
from pathlib import Path

import mysql.connector
import openpyxl

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("invias")

FOLDER = Path(os.environ.get("INVIAS_FOLDER", "."))
# Credenciales por variables de entorno (NO hardcodear secretos en el repo).
DB_REMOTE = {
    "host": os.environ.get("DB_HOST", ""),
    "port": int(os.environ.get("DB_PORT", "3306")),
    "database": os.environ.get("DB_NAME", "defaultdb"),
    "user": os.environ.get("DB_USER", ""),
    "password": os.environ.get("DB_PASSWORD", ""),
    "connect_timeout": 30,
    "ssl_ca": os.environ.get("DB_SSL_CA", "ca.pem"),
}
DB_LOCAL = {
    "host": os.environ.get("LOCAL_DB_HOST", "localhost"),
    "port": int(os.environ.get("LOCAL_DB_PORT", "3307")),
    "database": os.environ.get("LOCAL_DB_NAME", "interventoria"),
    "user": os.environ.get("LOCAL_DB_USER", "postgres"),
    "password": os.environ.get("LOCAL_DB_PASSWORD", "postgres"),
    "connect_timeout": 10,
}

CIUDAD_MAP = {
    ("AMAZONAS", "SELVA AMAZONICA"): "Leticia",
    ("ANTIOQUIA", "BAJO CAUCA"): "Caucasia",
    ("ANTIOQUIA", "MAGDALENA MEDIO"): "Puerto Berrio",
    ("ANTIOQUIA", "NORDESTE"): "Segovia",
    ("ANTIOQUIA", "NORTE"): "Yarumal",
    ("ANTIOQUIA", "OCCIDENTE"): "Santa Fe de Antioquia",
    ("ANTIOQUIA", "ORIENTE"): "Rionegro",
    ("ANTIOQUIA", "SUROESTE"): "Andes",
    ("ANTIOQUIA", "URABA"): "Apartado",
    ("ANTIOQUIA", "VALLE DE ABURRA"): "Medellin",
    ("ARAUCA", "ARAUCA"): "Arauca",
    ("ATLANTICO", "CENTRO ORIENTE"): "Barranquilla",
    ("ATLANTICO", "NORTE"): "Barranquilla",
    ("ATLANTICO", "OCCIDENTAL"): "Barranquilla",
    ("ATLANTICO", "SUR"): "Barranquilla",
    ("BOLIVAR", "DEPRESION MOMPOSINA"): "Santa Cruz de Mompox",
    ("BOLIVAR", "DIQUE BOLIVARENSE"): "Cartagena",
    ("BOLIVAR", "LOBA"): "San Pablo",
    ("BOLIVAR", "MAGDALENA MEDIO BOLIVARENSE"): "Simiti",
    ("BOLIVAR", "MOJANA BOLIVARENSE"): "Magangue",
    ("BOLIVAR", "MONTES DE MARIA"): "El Carmen de Bolivar",
    ("BOYACA", "CENTRO"): "Tunja",
    ("BOYACA", "GUTIERREZ"): "El Cocuy",
    ("BOYACA", "LA LIBERTAD"): "Labranzagrande",
    ("BOYACA", "LENGUPA"): "Miraflores",
    ("BOYACA", "MARQUEZ"): "Ramiriqui",
    ("BOYACA", "NEIRA"): "Garagoa",
    ("BOYACA", "NORTE"): "Soata",
    ("BOYACA", "OCCIDENTE"): "Muzo",
    ("BOYACA", "ORIENTE"): "Paez",
    ("BOYACA", "RICAURTE"): "Moniquira",
    ("BOYACA", "SUGAMUXI"): "Sogamoso",
    ("BOYACA", "TUNDAMA"): "Duitama",
    ("BOYACA", "VALDERRAMA"): "Soata",
    ("CALDAS", "ALTO OCCIDENTE"): "Supia",
    ("CALDAS", "ALTO ORIENTE"): "Pensilvania",
    ("CALDAS", "BAJO OCCIDENTE"): "La Dorada",
    ("CALDAS", "CENTRO"): "Manizales",
    ("CALDAS", "NORTE"): "Salamina",
    ("CALDAS", "ORIENTE"): "Pensilvania",
    ("CAQUETA", "CAQUETA"): "Florencia",
    ("CASANARE", "CASANARE"): "Yopal",
    ("CAUCA", "CENTRO"): "Popayan",
    ("CAUCA", "NORTE"): "Santander de Quilichao",
    ("CAUCA", "OCCIDENTE"): "Guapi",
    ("CAUCA", "ORIENTE"): "La Plata",
    ("CAUCA", "SUR"): "Bolivar",
    ("CESAR", "CENTRAL"): "Valledupar",
    ("CESAR", "NOROCCIDENTAL"): "La Paz",
    ("CESAR", "NORTE"): "Agustin Codazzi",
    ("CESAR", "SUR"): "Aguachica",
    ("CHOCO", "ATRATO"): "Quibdo",
    ("CHOCO", "DARIEN"): "Acandi",
    ("CHOCO", "PACIFICO NORTE"): "Nuqui",
    ("CHOCO", "PACIFICO SUR"): "Bahia Solano",
    ("CHOCO", "SAN JUAN"): "Istmina",
    ("CORDOBA", "ALTO SINU"): "Tierralta",
    ("CORDOBA", "BAJO SINU"): "Lorica",
    ("CORDOBA", "CENTRO"): "Monteria",
    ("CORDOBA", "COSTANERA"): "San Bernardo del Viento",
    ("CORDOBA", "SABANAS"): "San Antero",
    ("CORDOBA", "SAN JORGE"): "San Benito Abad",
    ("CORDOBA", "SINU MEDIO"): "Monteria",
    ("CUNDINAMARCA", "ALMEIDAS"): "Villapinzon",
    ("CUNDINAMARCA", "ALTO MAGDALENA"): "Girardot",
    ("CUNDINAMARCA", "BAJO MAGDALENA"): "Guaduas",
    ("CUNDINAMARCA", "GUALIVA"): "Villeta",
    ("CUNDINAMARCA", "GUAVIO"): "Gachala",
    ("CUNDINAMARCA", "MAGDALENA CENTRO"): "Pacho",
    ("CUNDINAMARCA", "MEDINA"): "Medina",
    ("CUNDINAMARCA", "ORIENTE"): "Fomeque",
    ("CUNDINAMARCA", "RIO NEGRO"): "La Calera",
    ("CUNDINAMARCA", "SABANA CENTRO"): "Zipaquira",
    ("CUNDINAMARCA", "SABANA OCCIDENTE"): "Facatativa",
    ("CUNDINAMARCA", "SOACHA"): "Soacha",
    ("CUNDINAMARCA", "SUMAPAZ"): "Fusagasuga",
    ("CUNDINAMARCA", "TEQUENDAMA"): "La Mesa",
    ("CUNDINAMARCA", "UBATE"): "Ubate",
    ("GUAINIA", "GUAINIA"): "Inirida",
    ("GUAVIARE", "GUAVIARE"): "San Jose del Guaviare",
    ("HUILA", "CENTRO"): "Neiva",
    ("HUILA", "NORTE"): "La Plata",
    ("HUILA", "OCCIDENTE"): "Pitalito",
    ("HUILA", "SUR"): "San Agustin",
    ("LA GUAJIRA", "NORTE"): "Riohacha",
    ("LA GUAJIRA", "SUR"): "Uribia",
    ("MAGDALENA", "CENTRO"): "Pivijay",
    ("MAGDALENA", "NORTE"): "Cienaga",
    ("MAGDALENA", "RIO"): "El Banco",
    ("MAGDALENA", "SANTA MARTA"): "Santa Marta",
    ("MAGDALENA", "SUR"): "Plato",
    ("META", "ARIARI"): "Granada",
    ("META", "CAPITAL"): "Villavicencio",
    ("META", "PIEDEMONTE"): "Acacias",
    ("META", "RIO META"): "Puerto Lopez",
    ("NARINO", "CENTRO"): "Pasto",
    ("NARINO", "CENTRO OCCIDENTE"): "Samaniego",
    ("NARINO", "COSTA"): "Tumaco",
    ("NARINO", "NORTE"): "Chachagui",
    ("NARINO", "SUR"): "Ipiales",
    ("NORTE DE SANTANDER", "CENTRO"): "Cucuta",
    ("NORTE DE SANTANDER", "NORTE"): "Pamplona",
    ("NORTE DE SANTANDER", "OCCIDENTE"): "Ocana",
    ("NORTE DE SANTANDER", "ORIENTE"): "Tibu",
    ("NORTE DE SANTANDER", "SUR OCCIDENTE"): "Chinacota",
    ("NORTE DE SANTANDER", "SUR ORIENTE"): "Cacota",
    ("PUTUMAYO", "PUTUMAYO"): "Mocoa",
    ("QUINDIO", "CAPITAL"): "Armenia",
    ("QUINDIO", "CORDILLERANOS"): "Salento",
    ("QUINDIO", "FRIA"): "Calarca",
    ("QUINDIO", "NORTE"): "Filandia",
    ("QUINDIO", "VALLE"): "La Tebaida",
    ("RISARALDA", "VERTIENTE DEL PACIFICO"): "Mistrato",
    ("RISARALDA", "VERTIENTE OCCIDENTE"): "Belen de Umbria",
    ("RISARALDA", "VERTIENTE ORIENTAL"): "Pereira",
    ("SAN ANDRES", "ARCHIPIELAGO DE SAN ANDRES"): "San Andres",
    ("SANTANDER", "COMUNERA"): "Socorro",
    ("SANTANDER", "GARCIA ROVIRA"): "Malaga",
    ("SANTANDER", "GUANENTA"): "San Gil",
    ("SANTANDER", "MARES"): "Barrancabermeja",
    ("SANTANDER", "SOTO"): "Bucaramanga",
    ("SANTANDER", "VELEZ"): "Velez",
    ("SUCRE", "LA MOJANA"): "Guaranda",
    ("SUCRE", "MONTES DE MARIA"): "Sincelejo",
    ("SUCRE", "MORROSQUILLO"): "Santiago de Tolu",
    ("SUCRE", "SABANAS"): "Corozal",
    ("SUCRE", "SAN JORGE"): "San Marcos",
    ("TOLIMA", "IBAGUE"): "Ibague",
    ("TOLIMA", "NEVADOS"): "Murillo",
    ("TOLIMA", "NORTE"): "Honda",
    ("TOLIMA", "ORIENTE"): "Melgar",
    ("TOLIMA", "SUR"): "Chaparral",
    ("TOLIMA", "SURORIENTE"): "Prado",
    ("VALLE DEL CAUCA", "CENTRO"): "Cali",
    ("VALLE DEL CAUCA", "NORTE"): "Cartago",
    ("VALLE DEL CAUCA", "OCCIDENTE"): "Buenaventura",
    ("VALLE DEL CAUCA", "ORIENTE"): "Tulua",
    ("VALLE DEL CAUCA", "SUR"): "Palmira",
    ("VAUPES", "VAUPES"): "Mitu",
    ("VICHADA", "VICHADA"): "Puerto Carreno",
}

INSERT_SQL = """INSERT IGNORE INTO apus
    (entidad, ciudad, pais, link_documento, numero_contrato,
     item, items_descripcion, item_unidad, precio_unitario,
     codigo_insumo, tipo_insumo, insumo_descripcion, insumo_unidad,
     rendimiento_insumo, precio_unitario_apu, precio_parcial_apu)
    VALUES (%s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s)"""


def parse_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "").replace(" ", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_filename(filename):
    m = re.match(r"APU_(\d+)_(.+?)__(.+?)_\d{4}_\d+\.xlsx", filename)
    if m:
        return m.group(1), m.group(2).replace("_", " ").strip(), m.group(3).replace("_", " ").strip()
    return None, None, None


import unicodedata

def strip_accents(s):
    return unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ascii")

def find_sheet(wb, name_like):
    target = strip_accents(name_like).upper().replace(" ", "").replace("'", "")
    for s in wb.sheetnames:
        candidate = strip_accents(s).upper().replace(" ", "").replace("'", "")
        if candidate == target:
            return s
    return None


TIPO_MAP = {"EQUIPO": "Equipos", "MATERIALES": "Materiales", "TRANSPORTE": "Transporte"}


def read_price_sheet(ws, code_col, desc_col, price_col, unit_col):
    prices = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or len(row) <= max(code_col, price_col, unit_col or 0):
            continue
        code = str(row[code_col] or "").strip()
        if not code:
            continue
        prices[code] = {
            "descripcion": str(row[desc_col] or "").strip(),
            "precio": parse_number(row[price_col]),
            "unidad": str(row[unit_col] or "").strip() if unit_col else "",
        }
    return prices


def read_labor_prices(ws):
    """Read MANO DE OBRA sheet. Region-specific. Returns {code: {desc, monthly, factor, workers}}."""
    labor = {}
    for row in ws.iter_rows(min_row=12, values_only=True):
        if not row or len(row) < 7:
            continue
        code = str(row[2] or "").strip() if len(row) > 2 else ""
        desc = str(row[3] or "").strip() if len(row) > 3 else ""
        monthly = parse_number(row[4]) if len(row) > 4 else None
        workers = parse_number(row[5]) if len(row) > 5 else None
        factor = parse_number(row[6]) if len(row) > 6 else None
        if code and monthly:
            labor[code] = {"descripcion": desc, "monthly": monthly, "factor": factor or 2.0, "workers": workers or 1}
    return labor


def extract_file(wb, dept, prov, filename):
    items = {}
    ciudad = CIUDAD_MAP.get((dept, prov), "")
    idx_name = find_sheet(wb, "INDICE")
    if idx_name:
        ws = wb[idx_name]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or len(row) < 12:
                continue
            item_code = str(row[4] or "").strip()
            if not item_code or item_code in ("ITEM DE PAGO", "#REF!"):
                continue
            items[item_code] = {
                "descripcion": str(row[5] or "").strip(),
                "unidad": str(row[6] or "").strip(),
                "costo_directo": parse_number(row[11]),
                "subtotal_equipos": parse_number(row[7]),
                "subtotal_materiales": parse_number(row[8]),
                "subtotal_transporte": parse_number(row[9]),
                "subtotal_mano_obra": parse_number(row[10]),
            }

    eq_name = find_sheet(wb, "EQUIPO")
    mat_name = find_sheet(wb, "MATERIALES")
    transp_name = find_sheet(wb, "TRANSPORTE")
    labor_name = find_sheet(wb, "MANO DE OBRA")

    equipos = read_price_sheet(wb[eq_name], code_col=2, desc_col=4, price_col=6, unit_col=3) if eq_name else {}
    materiales = read_price_sheet(wb[mat_name], code_col=2, desc_col=4, price_col=5, unit_col=3) if mat_name else {}
    transporte = read_price_sheet(wb[transp_name], code_col=2, desc_col=4, price_col=5, unit_col=3) if transp_name else {}
    labor = read_labor_prices(wb[labor_name]) if labor_name else {}

    apus_name = find_sheet(wb, "APUS")
    if not apus_name:
        log.warning("  APU sheet not found")
        return []
    ws = wb[apus_name]

    results = []
    basename = os.path.basename(filename)
    numero_contrato = f"INVIAS-2026-1-{filename.replace('.xlsx','').replace('APU_','')}"

    for ridx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        if not row:
            continue
        item_code = str(row[2] or "").strip() if len(row) > 2 else ""
        if not item_code or item_code in ("ITEM", "#", ""):
            continue

        item_desc = str(row[4] or "").strip() if len(row) > 4 else ""
        item_unidad = str(row[3] or "").strip() if len(row) > 3 else ""
        meta = items.get(item_code, {})
        costo_directo = meta.get("costo_directo")
        st_labor = meta.get("subtotal_mano_obra")

        hermeninv = str(row[5] or "").strip() if len(row) > 5 else ""
        pct_herramienta = parse_number(row[6]) if len(row) > 6 else None

        row_count = 0

        def add_row(codigo, tipo, descripcion, unidad, rendimiento, precio_unit):
            nonlocal row_count
            if precio_unit is not None and rendimiento is not None:
                parcial = round(rendimiento * precio_unit, 2)
                results.append((
                    "INSTITUTO NACIONAL DE VIAS", ciudad, "Colombia", basename, numero_contrato,
                    item_code, item_desc, item_unidad, costo_directo,
                    codigo, tipo, descripcion, unidad,
                    round(rendimiento, 10), round(precio_unit, 10), parcial,
                ))
                row_count += 1

        mae_cols = [7, 9, 11, 13, 15]
        for idx_mae, c_col in enumerate(mae_cols):
            q_col = c_col + 1
            if c_col < len(row) and q_col < len(row):
                code = str(row[c_col] or "").strip()
                qty = parse_number(row[q_col])
                if code and code != "HERMENINV" and qty is not None and qty > 0:
                    info = equipos.get(code, {})
                    add_row(code, "Equipos", info.get("descripcion", ""),
                            info.get("unidad", ""), qty, info.get("precio"))

        if hermeninv == "HERMENINV" and pct_herramienta and st_labor:
            add_row("HERMENINV", "Herramienta", "Herramienta menor (% Mano de obra)", "",
                    pct_herramienta, st_labor)

        for i in range(33, 65, 2):
            if i < len(row) and (i + 1) < len(row):
                code = str(row[i] or "").strip()
                qty = parse_number(row[i + 1])
                if code and qty is not None and qty > 0:
                    info = materiales.get(code, {})
                    add_row(code, "Materiales", info.get("descripcion", ""),
                            info.get("unidad", ""), qty, info.get("precio"))

        for i in range(65, 75, 2):
            if i < len(row) and (i + 1) < len(row):
                code = str(row[i] or "").strip()
                qty = parse_number(row[i + 1])
                if code and qty is not None and qty > 0:
                    info = transporte.get(code, {})
                    add_row(code, "Transporte", info.get("descripcion", ""),
                            info.get("unidad", ""), qty, info.get("precio"))

        for i in range(75, 95, 2):
            if i < len(row) and (i + 1) < len(row):
                code = str(row[i] or "").strip()
                rend = parse_number(row[i + 1])
                if code and rend is not None and rend > 0:
                    li = labor.get(code)
                    if li:
                        daily_jornal = li["monthly"] / 30.0
                        daily_cost = daily_jornal * li["factor"]
                        unit_price = daily_cost / rend
                        add_row(code, "Mano de obra", li["descripcion"], "", rend, unit_price)

        if row_count == 0 and costo_directo:
            add_row("SIN-DESGLOSE", "Indirectos", "Sin desglose disponible", "", 1, costo_directo)

    return results


def insert_batch(cursor, rows):
    try:
        cursor.executemany(INSERT_SQL, rows)
        affected = cursor.rowcount if cursor.rowcount >= 0 else len(rows)
        return affected
    except mysql.connector.Error as e:
        if "read-only" in str(e):
            raise
        log.warning("  DB batch error: %s", e)
        return 0


def get_db(local=False):
    cfg = DB_LOCAL if local else DB_REMOTE
    return mysql.connector.connect(**cfg)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--local", action="store_true", help="Use local MySQL (port 3307)")
    parser.add_argument("--retry", type=str, default="", help="Comma-separated filenames to retry")
    args = parser.parse_args()

    files = sorted(FOLDER.glob("APU_*.xlsx"))
    log.info("Found %d Excel files", len(files))
    if args.retry:
        retry_set = {f.strip() for f in args.retry.split(",")}
        files = [f for f in files if f.name in retry_set]
        log.info("Retrying %d files", len(files))
    elif args.limit:
        files = files[:args.limit]

    conn = None
    if not args.dry_run:
        conn = get_db(local=args.local)
        log.info("Connected to DB (%s)", "local" if args.local else "remote")

    total_rows = 0
    total_ok = 0
    total_err = 0

    for idx, fp in enumerate(files, 1):
        name = fp.name
        log.info("[%d/%d] %s", idx, len(files), name)
        code, dept, prov = parse_filename(name)
        if not dept:
            log.warning("  Could not parse filename, skipping")
            total_err += 1
            continue

        for attempt in range(3):
            try:
                wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
                t0 = time.time()
                rows = extract_file(wb, dept, prov, name)
                wb.close()
                elapsed = time.time() - t0
                log.info("  Extracted %d rows in %.1fs", len(rows), elapsed)

                if rows and conn and not args.dry_run:
                    cursor = conn.cursor()
                    inserted = insert_batch(cursor, rows)
                    conn.commit()
                    cursor.close()
                    total_rows += inserted
                    log.info("  Inserted %d (%d skipped)", inserted, len(rows) - inserted)
                elif rows:
                    total_rows += len(rows)
                total_ok += 1
                break
            except mysql.connector.Error as e:
                if "read-only" in str(e) and attempt < 2:
                    log.warning("  Read-only error, reconnecting and retrying...")
                    if conn:
                        try:
                            conn.close()
                        except Exception:
                            pass
                    time.sleep(2)
                    conn = get_db(local=args.local)
                    continue
                log.error("  DB ERROR (permanent): %s", e)
                total_err += 1
                break
            except Exception as e:
                log.error("  ERROR: %s", e)
                import traceback
                traceback.print_exc()
                total_err += 1
                break

    if conn:
        conn.close()

    log.info("=" * 50)
    log.info("SUMMARY: %d OK, %d errors, %d total rows", total_ok, total_err, total_rows)


if __name__ == "__main__":
    main()
