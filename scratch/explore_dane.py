import openpyxl

wb = openpyxl.load_workbook('scratch/dane_icoced_jul2026.xlsx', read_only=True, data_only=True)
print('Hojas:', wb.sheetnames)
for sn in wb.sheetnames[:8]:
    ws = wb[sn]
    print(f'\n=== {sn} (max_row={ws.max_row}, max_col={ws.max_column}) ===')
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
        vals = [str(c)[:40] if c is not None else '' for c in row]
        line = " | ".join(vals[:8])
        print(f'  R{i}: {line}')
