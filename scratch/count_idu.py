import openpyxl

wb = openpyxl.load_workbook('scratch/idu_2026.xlsx', read_only=True)
ws_apu = wb['APU']
ws_ins = wb['Inusmos']

count_apu = sum(1 for row in ws_apu.iter_rows(values_only=True) if any(row))
count_ins = sum(1 for row in ws_ins.iter_rows(values_only=True) if any(row))

print(f"Total APUs en IDU: {count_apu - 6}")
print(f"Total Insumos en IDU: {count_ins - 6}")
