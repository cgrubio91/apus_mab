import openpyxl

wb = openpyxl.load_workbook('scratch/idu_2026.xlsx', read_only=True)
print("Sheet names:", wb.sheetnames)

for sname in wb.sheetnames[:4]:
    print(f"\n--- Sheet: {sname} ---")
    ws = wb[sname]
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        if any(row):
            row_count += 1
            if row_count <= 8:
                # filter none
                filtered = [str(c)[:50] for c in row if c is not None]
                print(f"  Row {row_count}:", " | ".join(filtered[:8]))
            if row_count > 10:
                break
    print(f"Total non-empty sample rows seen: {row_count}")
