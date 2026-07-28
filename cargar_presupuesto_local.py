"""
Carga el presupuesto de Vías del Renacimiento a la BD local (Docker, localhost:3307).
"""

import re
import mysql.connector
from decimal import Decimal

import openpyxl

RUTA = r"C:\Users\cgrub\Downloads\invias 2026_1\Vias_del_Renacimiento_-_Acta_ACTA_30_generada_el_2026-07-27.xlsx"
HOJA = "ACTA_30_SegMAB"
NOMBRE_PROYECTO = "Vías del Renacimiento"

LOCAL_CONFIG = {
    "host": "localhost",
    "port": 3307,
    "database": "interventoria",
    "user": "postgres",
    "password": "postgres",
}


def normalizar(val):
    if val is None:
        return None
    s = str(val).strip().rstrip("_")
    return s if s else None


def es_capitulo(col_a):
    return col_a and re.match(r"^Capitulo\s+\d", str(col_a).strip(), re.IGNORECASE)


def es_resumen(col_a, col_e):
    a = (col_a or "").strip().upper()
    e = (col_e or "").strip().upper()
    keywords = ["SUBTOTAL", "REAJUSTE", "PLAN DE MANEJO", "PAGA", "AJUSTE A DISEÑOS", "TOTAL DE OBRA"]
    return any(k in a or k in e for k in keywords)


def main():
    wb = openpyxl.load_workbook(RUTA, data_only=True)
    ws = wb[HOJA]

    conn = mysql.connector.connect(**LOCAL_CONFIG)
    try:
        cur = conn.cursor(dictionary=True)

        # Borrar proyecto si ya existe (por id_folder)
        cur.execute("SELECT id FROM proyectos WHERE id_folder = 'vias_renacimiento_local'")
        existing = cur.fetchone()
        if existing:
            pid = existing["id"]
            cur.execute("DELETE FROM item_proyecto WHERE proyecto = %s", (pid,))
            cur.execute("DELETE FROM proyectos WHERE id = %s", (pid,))
            conn.commit()
            print("Proyecto anterior eliminado id=%d" % pid)

        # Obtener siguiente id_proy
        cur.execute("SELECT COALESCE(MAX(id_proy), 0) + 1 AS next_id FROM proyectos")
        next_id_proy = cur.fetchone()["next_id"]

        # Crear proyecto
        cur.execute(
            "INSERT INTO proyectos (id_proy, descripcion, id_folder) VALUES (%s, %s, %s)",
            (next_id_proy, NOMBRE_PROYECTO, "vias_renacimiento_local"),
        )
        proyecto_id = cur.lastrowid
        print("Proyecto creado: id=%d, id_proy=%d, desc=%s" % (proyecto_id, next_id_proy, NOMBRE_PROYECTO))

        # Parsear Excel
        chapter_id = None
        orden = 0

        for r in range(4, ws.max_row + 1):
            col_a = normalizar(ws.cell(r, 1).value)
            col_b = normalizar(ws.cell(r, 2).value)
            col_c = normalizar(ws.cell(r, 3).value)
            col_d = normalizar(ws.cell(r, 4).value)
            col_e = normalizar(ws.cell(r, 5).value)
            col_f = ws.cell(r, 6).value
            col_g = ws.cell(r, 7).value
            col_h = normalizar(ws.cell(r, 8).value)
            col_i = ws.cell(r, 9).value
            col_j = ws.cell(r, 10).value
            col_k = ws.cell(r, 11).value

            if not col_a and not col_e:
                continue
            if es_resumen(col_a, col_e):
                continue

            if es_capitulo(col_a):
                total = Decimal(str(col_k)) if col_k else Decimal(0)
                cur.execute(
                    """INSERT INTO item_proyecto
                       (proyecto, nivel, codigo, nombre, valor_presupuestado, tipo_item, orden, aprobado_interventoria)
                       VALUES (%s, 1, %s, %s, %s, 'PREVISTO', %s, 1)""",
                    (proyecto_id, col_a, col_a, total, orden),
                )
                chapter_id = cur.lastrowid
                orden += 1
                continue

            if chapter_id is None:
                continue

            cantidad = Decimal(str(col_i)) if col_i else Decimal(0)
            vr_unitario = Decimal(str(col_j)) if col_j else Decimal(0)
            vr_total = Decimal(str(col_k)) if col_k else Decimal(0)
            ai = Decimal(str(col_f)) if col_f else None
            aiu = Decimal(str(col_g)) if col_g else None

            cur.execute(
                """INSERT INTO item_proyecto
                   (proyecto, parent_id, nivel, codigo, especif_gral_raw, especif_part_raw,
                    grupo_ajuste_raw, nombre, unidad_medida, cantidad_presupuestada,
                    valor_unitario, valor_presupuestado, ai, aiu, tipo_item, orden, aprobado_interventoria)
                   VALUES (%s, %s, 2, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'PREVISTO', %s, 1)""",
                (proyecto_id, chapter_id, col_a, col_b, col_c, col_d,
                 col_e or col_a, col_h, cantidad, vr_unitario, vr_total,
                 ai, aiu, orden),
            )
            orden += 1

        # Actualizar presupuesto total
        cur.execute(
            "SELECT SUM(valor_presupuestado) AS total FROM item_proyecto WHERE proyecto = %s AND nivel = 1",
            (proyecto_id,),
        )
        total_presupuesto = cur.fetchone()["total"] or Decimal(0)
        cur.execute(
            "UPDATE proyectos SET presupuesto_total = %s WHERE id = %s",
            (total_presupuesto, proyecto_id),
        )

        conn.commit()
        print("Carga completada: proyecto_id=%d, total items=%d, presupuesto_total=%s" %
              (proyecto_id, orden, total_presupuesto))
        cur.close()

    except Exception:
        conn.rollback()
        print("Error durante la carga")
        raise
    finally:
        conn.close()
    wb.close()


if __name__ == "__main__":
    main()
