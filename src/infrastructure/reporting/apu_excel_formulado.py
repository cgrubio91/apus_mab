"""
Infrastructure: Exportador de APU en Excel con Fórmulas Vivas (openpyxl)
Genera plantillas de cálculo dinámicas donde los parciales, subtotales,
AIU y precio unitario final recalculan automáticamente si se alteran cantidades o precios.
"""

import io
from typing import Optional
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generar_apu_excel_formulado(solicitud: dict, desglose_aiu: dict, proyecto: Optional[dict] = None) -> bytes:
    """Genera un archivo Excel (.xlsx) con fórmulas matemáticas nativas de cálculo de APU."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análisis de Precio Unitario"

    # Estilos
    font_titulo = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    font_subtitulo = Font(name="Calibri", size=10, bold=True, color="334155")
    font_seccion = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=9, bold=True, color="1E293B")
    font_bold = Font(name="Calibri", size=9, bold=True)
    font_regular = Font(name="Calibri", size=9)

    fill_seccion = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_header = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_subtotal = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_total = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")  # amarillo suave

    thin_border = Side(style="thin", color="CBD5E1")
    border_box = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    double_bottom = Border(left=thin_border, right=thin_border, top=thin_border, bottom=Side(style="double", color="1E293B"))

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    num_format_currency = "$#,##0.00"
    num_format_rend = "0.0000"
    num_format_pct = "0.00%"

    # Encabezado
    ws["A1"] = "SISTEMA DE GESTIÓN Y AUDITORÍA DE PRECIOS UNITARIOS — MAPUS"
    ws["A1"].font = font_titulo
    ws["A2"] = "ANÁLISIS DE PRECIO UNITARIO (FORMATO FORMULADO AUDITABLE)"
    ws["A2"].font = font_subtitulo

    # Datos Generales
    nombre_proy = (proyecto.get("descripcion") if proyecto else None) or solicitud.get("nombre_proyecto") or "PROYECTO LOCAL"
    codigo_item = solicitud.get("codigo_item") or f"NPC-{solicitud.get('id', '')}"
    actividad = solicitud.get("descripcion_actividad") or "Ítem no previsto"
    unidad = solicitud.get("unidad_actividad") or "und"
    ciudad = solicitud.get("ciudad") or "Colombia"

    ws["A4"] = "PROYECTO:"
    ws["B4"] = nombre_proy
    ws["D4"] = "SOLICITUD #:"
    ws["E4"] = solicitud.get("id")

    ws["A5"] = "CÓDIGO ÍTEM:"
    ws["B5"] = codigo_item
    ws["D5"] = "UNIDAD:"
    ws["E5"] = unidad

    ws["A6"] = "ACTIVIDAD:"
    ws["B6"] = actividad
    ws["D6"] = "CIUDAD:"
    ws["E6"] = ciudad

    for r in range(4, 7):
        ws[f"A{r}"].font = font_bold
        ws[f"D{r}"].font = font_bold
        ws[f"B{r}"].font = font_regular
        ws[f"E{r}"].font = font_regular

    # Encabezado de la tabla de insumos
    fila = 8
    headers = ["CÓDIGO", "TIPO / DESCRIPCIÓN DEL INSUMO", "UNIDAD", "RENDIMIENTO / CANT.", "PRECIO UNITARIO ($)", "VALOR PARCIAL ($)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=fila, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_box
        cell.alignment = align_center if col_idx in (1, 3) else (align_right if col_idx >= 4 else align_left)

    insumos = solicitud.get("insumos") or []
    # Agrupar por categoría típica
    categorias = ["Materiales", "Mano de Obra", "Equipos", "Transporte / Otros"]
    subtotales_filas = []

    for cat in categorias:
        items_cat = [i for i in insumos if (i.get("tipo_insumo") or "").lower() == cat.lower() or
                     (cat == "Transporte / Otros" and (i.get("tipo_insumo") or "") not in ["Materiales", "Mano de Obra", "Equipos"])]
        if not items_cat:
            continue

        fila += 1
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
        cat_cell = ws.cell(row=fila, column=1, value=cat.upper())
        cat_cell.font = font_seccion
        cat_cell.fill = fill_seccion
        cat_cell.alignment = align_left

        inicio_cat = fila + 1
        for ins in items_cat:
            fila += 1
            cod = ins.get("codigo_insumo") or ""
            desc = ins.get("insumo_descripcion") or "Insumo"
            und = ins.get("insumo_unidad") or "und"
            rend = float(ins.get("rendimiento_insumo") or 1)
            p_u = float(ins.get("precio_unitario_apu") or ins.get("precio_banco") or 0)

            ws.cell(row=fila, column=1, value=cod).alignment = align_center
            ws.cell(row=fila, column=2, value=desc).alignment = align_left
            ws.cell(row=fila, column=3, value=und).alignment = align_center

            c_rend = ws.cell(row=fila, column=4, value=rend)
            c_rend.number_format = num_format_rend
            c_rend.alignment = align_right

            c_pu = ws.cell(row=fila, column=5, value=p_u)
            c_pu.number_format = num_format_currency
            c_pu.alignment = align_right

            # FÓRMULA VIVA DE PARCIAL: =D{fila}*E{fila}
            c_parc = ws.cell(row=fila, column=6, value=f"=D{fila}*E{fila}")
            c_parc.number_format = num_format_currency
            c_parc.alignment = align_right

            for c in range(1, 7):
                ws.cell(row=fila, column=c).border = border_box
                ws.cell(row=fila, column=c).font = font_regular

        fin_cat = fila
        # Fila de subtotal categoría
        fila += 1
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
        sub_label = ws.cell(row=fila, column=1, value=f"SUBTOTAL {cat.upper()}")
        sub_label.font = font_bold
        sub_label.alignment = align_right

        sub_val = ws.cell(row=fila, column=6, value=f"=SUM(F{inicio_cat}:F{fin_cat})")
        sub_val.font = font_bold
        sub_val.number_format = num_format_currency
        sub_val.alignment = align_right

        for c in range(1, 7):
            ws.cell(row=fila, column=c).fill = fill_subtotal
            ws.cell(row=fila, column=c).border = border_box

        subtotales_filas.append(fila)

    # TOTAL COSTO DIRECTO
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    cd_label = ws.cell(row=fila, column=1, value="TOTAL COSTO DIRECTO")
    cd_label.font = font_bold
    cd_label.alignment = align_right

    formula_cd = "+".join([f"F{f}" for f in subtotales_filas]) if subtotales_filas else "0"
    cd_val = ws.cell(row=fila, column=6, value=f"={formula_cd}")
    cd_val.font = font_bold
    cd_val.number_format = num_format_currency
    cd_val.alignment = align_right

    for c in range(1, 7):
        ws.cell(row=fila, column=c).fill = fill_subtotal
        ws.cell(row=fila, column=c).border = border_box

    fila_costo_directo = fila

    # ── TABLA DE A.I.U. FORMULADA ──
    fila += 2
    ws.cell(row=fila, column=1, value="CÁLCULO DE A.I.U. (COSTOS INDIRECTOS)").font = font_subtitulo

    pcts = desglose_aiu.get("porcentajes") or {}
    p_a = pcts.get("administracion", 15.0) / 100.0
    p_i = pcts.get("imprevistos", 3.0) / 100.0
    p_u = pcts.get("utilidad", 5.0) / 100.0
    p_iva = pcts.get("iva_utilidad", 19.0) / 100.0

    fila_aiu_header = fila + 1
    aiu_headers = ["CONCEPTO", "BASE", "%", "VALOR ($ COP)"]
    ws.cell(row=fila_aiu_header, column=1, value=aiu_headers[0]).font = font_header
    ws.cell(row=fila_aiu_header, column=2, value=aiu_headers[1]).font = font_header
    ws.cell(row=fila_aiu_header, column=3, value=aiu_headers[2]).font = font_header
    ws.cell(row=fila_aiu_header, column=4, value=aiu_headers[3]).font = font_header

    for c in range(1, 5):
        ws.cell(row=fila_aiu_header, column=c).fill = fill_header
        ws.cell(row=fila_aiu_header, column=c).border = border_box

    # Filas de A, I, U, IVA
    fila_a = fila_aiu_header + 1
    ws.cell(row=fila_a, column=1, value="Administración (A)").font = font_regular
    ws.cell(row=fila_a, column=2, value="Costo Directo").font = font_regular
    ws.cell(row=fila_a, column=3, value=p_a).number_format = num_format_pct
    ws.cell(row=fila_a, column=4, value=f"=F{fila_costo_directo}*C{fila_a}").number_format = num_format_currency

    fila_i = fila_a + 1
    ws.cell(row=fila_i, column=1, value="Imprevistos (I)").font = font_regular
    ws.cell(row=fila_i, column=2, value="Costo Directo").font = font_regular
    ws.cell(row=fila_i, column=3, value=p_i).number_format = num_format_pct
    ws.cell(row=fila_i, column=4, value=f"=F{fila_costo_directo}*C{fila_i}").number_format = num_format_currency

    fila_u = fila_i + 1
    ws.cell(row=fila_u, column=1, value="Utilidad (U)").font = font_regular
    ws.cell(row=fila_u, column=2, value="Costo Directo").font = font_regular
    ws.cell(row=fila_u, column=3, value=p_u).number_format = num_format_pct
    ws.cell(row=fila_u, column=4, value=f"=F{fila_costo_directo}*C{fila_u}").number_format = num_format_currency

    fila_iva = fila_u + 1
    ws.cell(row=fila_iva, column=1, value="IVA sobre Utilidad").font = font_regular
    ws.cell(row=fila_iva, column=2, value="Utilidad (U)").font = font_regular
    ws.cell(row=fila_iva, column=3, value=p_iva).number_format = num_format_pct
    ws.cell(row=fila_iva, column=4, value=f"=D{fila_u}*C{fila_iva}").number_format = num_format_currency

    for f in range(fila_a, fila_iva + 1):
        for c in range(1, 5):
            ws.cell(row=f, column=c).border = border_box

    fila_total_aiu = fila_iva + 1
    ws.merge_cells(start_row=fila_total_aiu, start_column=1, end_row=fila_total_aiu, end_column=3)
    ws.cell(row=fila_total_aiu, column=1, value="TOTAL A.I.U.").font = font_bold
    ws.cell(row=fila_total_aiu, column=1).alignment = align_right
    ws.cell(row=fila_total_aiu, column=4, value=f"=SUM(D{fila_a}:D{fila_iva})").number_format = num_format_currency
    ws.cell(row=fila_total_aiu, column=4).font = font_bold

    for c in range(1, 5):
        ws.cell(row=fila_total_aiu, column=c).fill = fill_subtotal
        ws.cell(row=fila_total_aiu, column=c).border = border_box

    # TOTAL PRECIO UNITARIO
    fila_total_final = fila_total_aiu + 2
    ws.merge_cells(start_row=fila_total_final, start_column=1, end_row=fila_total_final, end_column=3)
    lbl_final = ws.cell(row=fila_total_final, column=1, value="PRECIO UNITARIO TOTAL (COSTO DIRECTO + A.I.U.)")
    lbl_final.font = font_bold
    lbl_final.alignment = align_right

    val_final = ws.cell(row=fila_total_final, column=4, value=f"=F{fila_costo_directo}+D{fila_total_aiu}")
    val_final.font = font_bold
    val_final.number_format = num_format_currency
    val_final.alignment = align_right

    for c in range(1, 5):
        ws.cell(row=fila_total_final, column=c).fill = fill_total
        ws.cell(row=fila_total_final, column=c).border = double_bottom

    # Ajustar anchos de columna automáticos
    col_widths = {1: 14, 2: 44, 3: 10, 4: 22, 5: 22, 6: 22}
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
