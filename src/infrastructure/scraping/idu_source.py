"""
Infrastructure: Adaptador y Scraper de Precios de Referencia del IDU (Bogotá)

Inspecciona el portal oficial de SIIPVIALES - Componente Económico del IDU
(https://www.idu.gov.co/page/siipviales/economico/portafolio), detecta la última
versión de la base de precios de referencia (Visor_BPR_*.xlsx) y parsea tanto
los APUs completos como los Insumos desglosados en entidades ReferenciaExterna.
"""

import logging
import os
import re
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl
import requests

from src.domain.entities.referencia_externa import ReferenciaExterna

log = logging.getLogger("mapus.infrastructure.idu")

FUENTE = "IDU"
PORTAFOLIO_URL = "https://www.idu.gov.co/page/siipviales/economico/portafolio"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


class IduSource:
    """Crawler y parser del Sistema de Información de Precios de Referencia del IDU."""

    def __init__(self, timeout: int = 15):
        self.timeout = timeout

    def obtener_url_ultimo_excel(self) -> Optional[str]:
        """Consulta el portal del IDU y encuentra el enlace a la última matriz de precios."""
        try:
            r = requests.get(PORTAFOLIO_URL, headers=HEADERS, timeout=self.timeout)
            if r.status_code != 200:
                log.warning("No se pudo acceder al portal del IDU: HTTP %d", r.status_code)
                return None
            
            links = re.findall(r'href="([^"]+)"', r.text)
            candidatos = [
                l for l in links
                if "visor_bpr" in l.lower() and l.lower().endswith(".xlsx")
            ]
            if not candidatos:
                # Búsqueda fallback de cualquier archivo de precios reciente
                candidatos = [
                    l for l in links
                    if "precios_unitarios" in l.lower() and l.lower().endswith(".xlsx")
                ]

            if candidatos:
                # El primero suele ser el más reciente publicado
                url = candidatos[0]
                if not url.startswith("http"):
                    url = f"https://www.idu.gov.co{url}"
                log.info("Última matriz IDU detectada: %s", url)
                return url
        except Exception as e:
            log.warning("Error rastreando portal del IDU: %s", e)
        return None

    def descargar_excel(self, ruta_destino: str = "scratch/idu_2026.xlsx") -> str:
        """Descarga el último Excel del IDU si no existe localmente."""
        destino = Path(ruta_destino)
        if destino.exists() and destino.stat().st_size > 100_000:
            log.info("Archivo IDU ya existe en caché: %s (%d bytes)", destino, destino.stat().st_size)
            return str(destino)

        url = self.obtener_url_ultimo_excel()
        if not url:
            # Fallback a URL conocida de 2026-I Fase I si la web no responde
            url = "https://www.idu.gov.co/Archivos_Portal/2026/transparencia/inf-de-interes/siipviales/economico/07-julio/1A.%20Visor_BPR_2026-I_Fase_I_29-07-2026.xlsx"

        log.info("Descargando matriz IDU desde: %s", url)
        destino.parent.mkdir(parents=True, exist_ok=True)
        r = requests.get(url, headers=HEADERS, timeout=30, stream=True)
        r.raise_for_status()

        with open(destino, "wb") as f:
            for chunk in r.iter_content(chunk_size=65536):
                f.write(chunk)

        log.info("Descarga IDU completada en %s (%d bytes)", destino, destino.stat().st_size)
        return str(destino)

    def parsear_excel(self, ruta_excel: str) -> tuple[list[ReferenciaExterna], list[ReferenciaExterna]]:
        """Lee las hojas 'APU' e 'Inusmos' del Excel oficial del IDU y genera entidades de dominio."""
        wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)
        apus: list[ReferenciaExterna] = []
        insumos: list[ReferenciaExterna] = []

        fecha_vigencia = date(2026, 7, 29)

        # ── 1. Hoja APU ──
        if "APU" in wb.sheetnames:
            ws_apu = wb["APU"]
            header_found = False
            col_map = {}

            for row_idx, row in enumerate(ws_apu.iter_rows(values_only=True)):
                if not row or not any(row):
                    continue

                if not header_found:
                    str_row = [str(c).strip().lower() for c in row if c is not None]
                    if any("código" in s or "codigo" in s for s in str_row) and any("nombre" in s for s in str_row):
                        header_found = True
                        for idx, val in enumerate(row):
                            if val is not None:
                                s = str(val).strip().lower()
                                col_map[s] = idx
                        continue
                else:
                    # Fila de datos
                    try:
                        c_nom = col_map.get("nombre")
                        c_val = col_map.get("valor", col_map.get("precio"))
                        c_cod = col_map.get("código", col_map.get("codigo"))
                        c_um = col_map.get("um", col_map.get("unidad"))
                        c_cap = col_map.get("capítulo", col_map.get("capitulo"))
                        c_sub = col_map.get("subcapítulo", col_map.get("subcapitulo"))
                        c_fec = col_map.get("fecha de actualización", col_map.get("fecha"))

                        nombre = str(row[c_nom]).strip() if c_nom is not None and row[c_nom] is not None else None
                        valor_raw = row[c_val] if c_val is not None else None
                        if not nombre or valor_raw is None:
                            continue

                        try:
                            precio = Decimal(str(valor_raw).replace(",", "."))
                        except (InvalidOperation, TypeError, ValueError):
                            continue

                        codigo = str(row[c_cod]).strip() if c_cod is not None and row[c_cod] is not None else None
                        unidad = str(row[c_um]).strip() if c_um is not None and row[c_um] is not None else "und"
                        capitulo = str(row[c_cap]).strip() if c_cap is not None and row[c_cap] is not None else ""
                        subcapitulo = str(row[c_sub]).strip() if c_sub is not None and row[c_sub] is not None else ""
                        fecha_str = str(row[c_fec]).strip() if c_fec is not None and row[c_fec] is not None else "2026-I"

                        observacion = f"{capitulo} > {subcapitulo} · {fecha_str}".strip(" >· ")

                        apus.append(ReferenciaExterna(
                            fuente=FUENTE,
                            fuente_id=f"APU-{codigo or row_idx}",
                            url=PORTAFOLIO_URL,
                            granularidad="item",
                            descripcion=nombre,
                            unidad=unidad,
                            codigo=codigo,
                            precio=precio,
                            rendimiento=None,
                            ciudad="Bogota",
                            departamento="Bogota D.C.",
                            entidad="Instituto de Desarrollo Urbano - IDU",
                            proveedor=None,
                            fecha=fecha_vigencia,
                            observacion=observacion,
                        ))
                    except Exception as e:
                        log.debug("Error fila APU %d: %s", row_idx, e)
                        continue

        # ── 2. Hoja Inusmos (Insumos) ──
        target_ins = "Inusmos" if "Inusmos" in wb.sheetnames else ("Insumos" if "Insumos" in wb.sheetnames else None)
        if target_ins:
            ws_ins = wb[target_ins]
            header_found = False
            col_map = {}

            for row_idx, row in enumerate(ws_ins.iter_rows(values_only=True)):
                if not row or not any(row):
                    continue

                if not header_found:
                    str_row = [str(c).strip().lower() for c in row if c is not None]
                    if any("código" in s or "codigo" in s for s in str_row) and any("precio" in s for s in str_row):
                        header_found = True
                        for idx, val in enumerate(row):
                            if val is not None:
                                s = str(val).strip().lower()
                                col_map[s] = idx
                        continue
                else:
                    # Fila de datos
                    try:
                        c_nom = col_map.get("nombre")
                        c_val = col_map.get("precio", col_map.get("valor"))
                        c_cod = col_map.get("código", col_map.get("codigo"))
                        c_um = col_map.get("um", col_map.get("unidad"))
                        c_grp = col_map.get("grupo de base de datos", col_map.get("grupo"))
                        c_fec = col_map.get("fecha de actualización", col_map.get("fecha"))

                        nombre = str(row[c_nom]).strip() if c_nom is not None and row[c_nom] is not None else None
                        precio_raw = row[c_val] if c_val is not None else None
                        if not nombre or precio_raw is None:
                            continue

                        try:
                            precio = Decimal(str(precio_raw).replace(",", "."))
                        except (InvalidOperation, TypeError, ValueError):
                            continue

                        codigo = str(row[c_cod]).strip() if c_cod is not None and row[c_cod] is not None else None
                        unidad = str(row[c_um]).strip() if c_um is not None and row[c_um] is not None else "und"
                        grupo = str(row[c_grp]).strip() if c_grp is not None and row[c_grp] is not None else ""
                        fecha_str = str(row[c_fec]).strip() if c_fec is not None and row[c_fec] is not None else "2026-I"
                        observacion = f"Insumo: {grupo} · {fecha_str}".strip(" · ")

                        insumos.append(ReferenciaExterna(
                            fuente=FUENTE,
                            fuente_id=f"INS-{codigo or row_idx}",
                            url=PORTAFOLIO_URL,
                            granularidad="insumo",
                            descripcion=nombre,
                            unidad=unidad,
                            codigo=codigo,
                            precio=precio,
                            rendimiento=None,
                            ciudad="Bogota",
                            departamento="Bogota D.C.",
                            entidad="Instituto de Desarrollo Urbano - IDU",
                            proveedor=None,
                            fecha=fecha_vigencia,
                            observacion=observacion,
                        ))
                    except Exception as e:
                        log.debug("Error fila Insumo %d: %s", row_idx, e)
                        continue

        wb.close()
        log.info("Parsea IDU completado: %d APUs, %d Insumos", len(apus), len(insumos))
        return apus, insumos
