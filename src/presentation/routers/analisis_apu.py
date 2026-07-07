"""
Presentation: Análisis APU Routes
Approval workflow endpoints.
"""

import logging
from typing import Optional, List

from fastapi import APIRouter, HTTPException, status, UploadFile, File, Depends
from src.domain.entities.analisis import AnalisisApuCreate, RechazarRequest
from src.application.use_cases.manage_analisis import (
    crear_solicitud,
    get_solicitudes,
    get_solicitud,
    realizar_analisis,
    preaprobar,
    rechazar,
    nuevas_cotizaciones_recibidas,
    aprobar_subgerente,
    firmar_legal,
    get_aprendizaje_rechazos,
)
from src.presentation.auth import get_current_user, require_role, get_optional_user

log = logging.getLogger("mapus.presentation.analisis")
router = APIRouter()


@router.post("/analisis-apu/upload", tags=["Análisis APU"])
async def upload_cotizacion(files: List[UploadFile] = File(...)) -> dict:
    try:
        grupos_insumos = []

        for idx, file in enumerate(files):
            filename = file.filename or f"archivo_{idx+1}"
            ext = filename.split(".")[-1].lower() if "." in filename else ""

            content = await file.read()
            if len(content) > 50 * 1024 * 1024:
                raise HTTPException(status_code=413, detail=f"{filename}: Archivo demasiado grande (máx 50 MB)")

            raw_insumos = []
            if ext == "pdf":
                import tempfile
                from src.infrastructure.ai.gemini_extractor import extract_apus_from_pdf_batched

                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                    tmp.write(content)
                    tmp_path = tmp.name
                try:
                    raw_insumos = extract_apus_from_pdf_batched(tmp_path, filename)
                finally:
                    import os
                    if os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except PermissionError:
                            pass
            elif ext in ("xlsx", "xls"):
                import tempfile
                from src.infrastructure.ai.gemini_extractor import extract_apus_from_excel

                with tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}") as tmp:
                    tmp.write(content)
                    tmp_path = tmp.name
                try:
                    raw_insumos = extract_apus_from_excel(tmp_path, filename)
                finally:
                    import os
                    if os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except PermissionError:
                            pass
            else:
                raise HTTPException(status_code=400, detail=f"{filename}: Formato no soportado. Use PDF o Excel.")

            from src.infrastructure.ai.gemini_extractor import post_process_extracted_data
            insumos = post_process_extracted_data(raw_insumos, filename)

            if insumos:
                grupos_insumos.append({"grupo_cotizacion": idx + 1, "nombre_archivo": filename, "insumos": insumos})

        if not grupos_insumos:
            raise HTTPException(status_code=400, detail="No se pudieron extraer insumos de los archivos.")

        total_insumos = sum(len(g["insumos"]) for g in grupos_insumos)
        solicitud_id = crear_solicitud(grupos_insumos)

        return {
            "success": True,
            "solicitud_id": solicitud_id,
            "archivos_procesados": len(grupos_insumos),
            "total_insumos": total_insumos,
            "grupos": [{"grupo": g["grupo_cotizacion"], "archivo": g["nombre_archivo"], "insumos": len(g["insumos"])} for g in grupos_insumos],
            "mensaje": f"Solicitud #{solicitud_id} creada con {len(grupos_insumos)} archivo(s) y {total_insumos} ítems.",
        }

    except HTTPException:
        raise
    except Exception as e:
        log.exception("Error en upload cotización")
        raise HTTPException(status_code=500, detail="Error procesando archivo(s). Revisa los logs para más detalle.")


@router.post("/analisis-apu/crear", tags=["Análisis APU"])
async def crear_solicitud_manual(payload: AnalisisApuCreate) -> dict:
    try:
        grupo = {"grupo_cotizacion": 1, "nombre_archivo": payload.link_documento, "insumos": [i.model_dump(exclude={"id", "solicitud_id"}) for i in payload.insumos]}
        solicitud_id = crear_solicitud([grupo])
        return {"success": True, "solicitud_id": solicitud_id}
    except Exception as e:
        log.exception("Error creando solicitud manual")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.post("/analisis-apu/{solicitud_id}/analizar", tags=["Análisis APU"])
async def analizar_solicitud(solicitud_id: int) -> dict:
    try:
        resultado = realizar_analisis(solicitud_id)
        return {"success": True, **resultado}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log.exception("Error analizando solicitud")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.get("/analisis-apu", tags=["Análisis APU"])
async def listar_solicitudes(estado: Optional[str] = None) -> dict:
    try:
        solicitudes = get_solicitudes(estado)
        for s in solicitudes:
            for campo in ("fecha_solicitud", "fecha_limite_respuesta", "fecha_limite_aprobacion", "created_at", "updated_at"):
                if s.get(campo) and hasattr(s[campo], "isoformat"):
                    s[campo] = s[campo].isoformat()
        return {"success": True, "data": solicitudes}
    except Exception as e:
        log.exception("Error listando solicitudes")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.get("/analisis-apu/{solicitud_id}", tags=["Análisis APU"])
async def obtener_solicitud(solicitud_id: int) -> dict:
    try:
        solicitud = get_solicitud(solicitud_id)
        if not solicitud:
            raise HTTPException(status_code=404, detail="Solicitud no encontrada")

        for campo in ("fecha_solicitud", "fecha_limite_respuesta", "fecha_limite_aprobacion", "created_at", "updated_at"):
            if solicitud.get(campo) and hasattr(solicitud[campo], "isoformat"):
                solicitud[campo] = solicitud[campo].isoformat()

        if solicitud.get("analisis"):
            if solicitud["analisis"].get("created_at") and hasattr(solicitud["analisis"]["created_at"], "isoformat"):
                solicitud["analisis"]["created_at"] = solicitud["analisis"]["created_at"].isoformat()

        for h in solicitud.get("historial", []):
            if h.get("created_at") and hasattr(h["created_at"], "isoformat"):
                h["created_at"] = h["created_at"].isoformat()

        return {"success": True, "data": solicitud}
    except HTTPException:
        raise
    except Exception as e:
        log.exception("Error obteniendo solicitud")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.post("/analisis-apu/{solicitud_id}/preaprobar", tags=["Análisis APU"])
async def preaprobar_apu(solicitud_id: int, user: dict = Depends(require_role("analista"))) -> dict:
    try:
        return preaprobar(solicitud_id, user["rol"], user["nombre"])
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log.exception("Error en preaprobación")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.post("/analisis-apu/{solicitud_id}/rechazar", tags=["Análisis APU"])
async def rechazar_apu(solicitud_id: int, payload: RechazarRequest, user: dict = Depends(require_role("analista"))) -> dict:
    try:
        return rechazar(solicitud_id, user["rol"], user["nombre"], payload.motivo)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log.exception("Error en rechazo")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.post("/analisis-apu/{solicitud_id}/nuevas-cotizaciones", tags=["Análisis APU"])
async def nuevas_cotizaciones(solicitud_id: int) -> dict:
    try:
        return nuevas_cotizaciones_recibidas(solicitud_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log.exception("Error registrando nuevas cotizaciones")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.post("/analisis-apu/{solicitud_id}/aprobar-subgerente", tags=["Análisis APU"])
async def aprobar_subgerente_endpoint(solicitud_id: int, user: dict = Depends(require_role("subgerente"))) -> dict:
    try:
        return aprobar_subgerente(solicitud_id, user["rol"], user["nombre"])
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log.exception("Error en aprobación de subgerente")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.post("/analisis-apu/{solicitud_id}/firmar-legal", tags=["Análisis APU"])
async def firmar_legal_endpoint(solicitud_id: int, user: dict = Depends(require_role("legal"))) -> dict:
    try:
        return firmar_legal(solicitud_id, user["rol"], user["nombre"])
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log.exception("Error en firma legal")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.get("/analisis-apu/{solicitud_id}/export", tags=["Análisis APU"])
async def exportar_analisis(solicitud_id: int):
    """Exporta a Excel el análisis comparativo de una solicitud (ítems + resumen)."""
    import io
    from datetime import datetime
    from fastapi.responses import StreamingResponse

    try:
        solicitud = get_solicitud(solicitud_id)
        if not solicitud:
            raise HTTPException(status_code=404, detail="Solicitud no encontrada")
        analisis = solicitud.get("analisis") or {}
        items = analisis.get("items_analizados") or []
        if not items:
            raise HTTPException(status_code=400, detail="La solicitud aún no tiene análisis para exportar")

        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        columnas = [
            ("grupo_cotizacion", "COTIZACIÓN"),
            ("item", "ITEM"),
            ("descripcion", "DESCRIPCIÓN"),
            ("unidad", "UNIDAD"),
            ("precio_ofertado", "PRECIO OFERTADO"),
            ("mejor_precio_banco", "MEJOR PRECIO BANCO"),
            ("diferencia_precio", "DIFERENCIA"),
            ("existe_en_banco", "EN BANCO"),
            ("estructura_insumos_coincide", "ESTRUCTURA OK"),
            ("rendimiento_coincide", "RENDIMIENTO OK"),
            ("recomendacion", "RECOMENDACIÓN IA"),
            ("observaciones", "OBSERVACIONES"),
        ]

        def build() -> bytes:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Análisis #{solicitud_id}"
            ws.append([f"Solicitud #{solicitud_id} — {solicitud.get('nombre_proyecto') or ''}"])
            ws.append([f"Estado: {solicitud.get('estado')} · Recomendación IA: {analisis.get('recomendacion') or 'N/A'}"])
            ws.append([f"Resumen: {analisis.get('resumen') or ''}"])
            ws.append([])
            ws.append([label for _, label in columnas])
            for it in items:
                fila = []
                for key, _ in columnas:
                    v = it.get(key)
                    if isinstance(v, bool):
                        v = "Sí" if v else "No"
                    fila.append(v if v is not None else "")
                ws.append(fila)
            for i, (_, label) in enumerate(columnas, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(14, min(50, len(label) + 6))
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        import asyncio
        content = await asyncio.to_thread(build)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="analisis_solicitud_{solicitud_id}_{stamp}.xlsx"'},
        )
    except HTTPException:
        raise
    except Exception:
        log.exception("Error exportando análisis %s", solicitud_id)
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")


@router.get("/analisis-apu/aprendizaje/rechazos", tags=["Análisis APU"])
async def obtener_aprendizaje(limit: int = 20) -> dict:
    try:
        data = get_aprendizaje_rechazos(limit)
        return {"success": True, "data": data}
    except Exception as e:
        log.exception("Error obteniendo aprendizaje")
        raise HTTPException(status_code=500, detail="Error interno del servidor. Revisa los logs para más detalle.")
