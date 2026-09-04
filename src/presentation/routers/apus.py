"""
Presentation: APU Query Routes
"""

import asyncio
import csv
import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Query, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
import mysql.connector
from pydantic import BaseModel

from src.infrastructure.database.connection import execute_query

from src.application.use_cases.query_apus import (
    get_apus as query_apus,
    get_filter_options,
    get_dashboard_stats,
    get_unique_projects,
    delete_project,
    ALLOWED_SORT_FIELDS,
    MAX_LIMIT,
)
from src.presentation.auth import require_role

log = logging.getLogger("mapus.presentation.apus")
router = APIRouter()


class ApuQueryFilters(BaseModel):
    nombre_proyecto: Optional[str] = None
    ciudad: Optional[str] = None
    items_descripcion: Optional[str] = None
    insumo_descripcion: Optional[str] = None
    tipo_insumo: Optional[str] = None
    contratista: Optional[str] = None
    entidad: Optional[str] = None
    codigo_insumo: Optional[str] = None
    item: Optional[str] = None
    item_unidad: Optional[str] = None
    insumo_unidad: Optional[str] = None
    pais: Optional[str] = None
    numero_contrato: Optional[str] = None


@router.get("/apus", tags=["APUs"])
async def get_apus_endpoint(
    filters_params: ApuQueryFilters = Depends(),
    search: Optional[str] = Query(None, pattern=r"^[a-zA-Z0-9\s\-_.,\+áéíóúÁÉÍÓÚñÑ]+$"),
    sort_by: Optional[str] = Query(None, pattern=f"^({'|'.join(ALLOWED_SORT_FIELDS)})$"),
    sort_order: str = Query("asc", pattern="^(asc|desc)$"),
    limit: int = Query(50, ge=1, le=MAX_LIMIT),
    offset: int = Query(0, ge=0),
) -> dict:
    try:
        filters = {k: v for k, v in filters_params.model_dump().items() if v is not None}
        result = await asyncio.to_thread(query_apus, filters, limit, offset, sort_by=sort_by, sort_order=sort_order, search=search)
        return result
    except mysql.connector.Error:
        log.exception("Database error in get_apus")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al procesar la consulta en la base de datos.")
    except Exception:
        log.exception("Critical error in get_apus_endpoint")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Ocurrió un error interno al recuperar los registros.")


EXPORT_MAX_ROWS = 10000
EXPORT_COLUMNS = [
    "id", "nombre_proyecto", "ciudad", "pais", "entidad", "contratista", "numero_contrato",
    "fecha_aprobacion_apu", "fecha_analisis_apu", "item", "items_descripcion", "item_unidad",
    "precio_unitario", "precio_unitario_sin_aiu", "codigo_insumo", "tipo_insumo",
    "insumo_descripcion", "insumo_unidad", "rendimiento_insumo", "precio_unitario_apu",
    "precio_parcial_apu", "observacion", "link_documento",
]


def _fetch_all_for_export(filters: dict, search: Optional[str], sort_by: Optional[str], sort_order: str) -> list[dict]:
    """Recorre el banco por páginas (respetando MAX_LIMIT del repo) hasta EXPORT_MAX_ROWS."""
    rows: list[dict] = []
    offset = 0
    while len(rows) < EXPORT_MAX_ROWS:
        page = query_apus(filters, MAX_LIMIT, offset, sort_by=sort_by, sort_order=sort_order, search=search)
        data = page.get("data", [])
        rows.extend(data)
        if len(data) < MAX_LIMIT:
            break
        offset += MAX_LIMIT
    return rows[:EXPORT_MAX_ROWS]


def _rows_to_xlsx(rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Banco de APUs"
    ws.append([c.upper() for c in EXPORT_COLUMNS])
    for row in rows:
        ws.append([_cell(row.get(c)) for c in EXPORT_COLUMNS])
    for i, col in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(col) + 4))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(value):
    # openpyxl no acepta objetos arbitrarios: normaliza fechas/decimales a tipos simples
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    from decimal import Decimal
    if isinstance(value, Decimal):
        return float(value)
    return value


def _rows_to_csv(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([c.upper() for c in EXPORT_COLUMNS])
    for row in rows:
        writer.writerow([_cell(row.get(c)) for c in EXPORT_COLUMNS])
    # BOM para que Excel abra el CSV con acentos correctos
    return ("﻿" + buf.getvalue()).encode("utf-8")


@router.get("/apus/export", tags=["APUs"])
async def export_apus(
    filters_params: ApuQueryFilters = Depends(),
    search: Optional[str] = Query(None, pattern=r"^[a-zA-Z0-9\s\-_.,\+áéíóúÁÉÍÓÚñÑ]+$"),
    sort_by: Optional[str] = Query(None, pattern=f"^({'|'.join(ALLOWED_SORT_FIELDS)})$"),
    sort_order: str = Query("asc", pattern="^(asc|desc)$"),
    formato: str = Query("xlsx", pattern="^(xlsx|csv)$"),
):
    try:
        filters = {k: v for k, v in filters_params.model_dump().items() if v is not None}
        rows = await asyncio.to_thread(_fetch_all_for_export, filters, search, sort_by, sort_order)

        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        if formato == "csv":
            content = await asyncio.to_thread(_rows_to_csv, rows)
            media = "text/csv; charset=utf-8"
            filename = f"banco_apus_{stamp}.csv"
        else:
            content = await asyncio.to_thread(_rows_to_xlsx, rows)
            media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"banco_apus_{stamp}.xlsx"

        return StreamingResponse(
            io.BytesIO(content),
            media_type=media,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except mysql.connector.Error:
        log.exception("Database error in export_apus")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al exportar los registros.")
    except Exception:
        log.exception("Critical error in export_apus")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Ocurrió un error interno al exportar.")


@router.get("/apus/filter-options", tags=["APUs"])
async def get_apus_filter_options() -> dict:
    try:
        return get_filter_options()
    except mysql.connector.Error:
        log.exception("Database error in get_apus_filter_options")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al cargar las opciones de filtro.")


@router.get("/apus/historico-precios", tags=["APUs"])
async def historico_precios(
    insumo: str = Query(..., min_length=3, max_length=200),
    ciudad: Optional[str] = Query(None, max_length=100),
    nombre_proyecto: Optional[str] = Query(None, max_length=200),
) -> dict:
    """Evolución mensual del precio de un insumo (avg/min/max) para graficar tendencias."""
    from src.infrastructure.database.connection import execute_query

    try:
        # Búsqueda inteligente por palabras clave (igual que Análisis APU): encuentra
        # las descripciones del banco más parecidas y agrega sus precios por mes.
        from src.infrastructure.database.repositories.analisis_repository import analisis_repo
        candidatos = await asyncio.to_thread(analisis_repo.buscar_insumos_candidatos, insumo, 25)
        descripciones = [c["insumo_descripcion"] for c in candidatos if c.get("insumo_descripcion")]

        condiciones = ["precio_unitario_apu IS NOT NULL", "precio_unitario_apu > 0"]
        params: list = []
        if descripciones:
            marcadores = ",".join(["%s"] * len(descripciones))
            condiciones.append(f"insumo_descripcion IN ({marcadores})")
            params.extend(descripciones)
        else:
            condiciones.append("insumo_descripcion LIKE %s")
            params.append(f"%{insumo}%")
        if ciudad:
            condiciones.append("TRIM(ciudad) = TRIM(%s)")
            params.append(ciudad)
        if nombre_proyecto:
            condiciones.append("TRIM(nombre_proyecto) = TRIM(%s)")
            params.append(nombre_proyecto)

        rows = await asyncio.to_thread(
            execute_query,
            f"""SELECT CONCAT(
                       YEAR(COALESCE(fecha_analisis_apu, fecha_aprobacion_apu, created_at)), '-',
                       LPAD(MONTH(COALESCE(fecha_analisis_apu, fecha_aprobacion_apu, created_at)), 2, '0')
                   ) AS periodo,
                       AVG(precio_unitario_apu) AS precio_promedio,
                       MIN(precio_unitario_apu) AS precio_minimo,
                       MAX(precio_unitario_apu) AS precio_maximo,
                       COUNT(*) AS registros
                FROM apus
                WHERE {' AND '.join(condiciones)}
                GROUP BY periodo
                ORDER BY periodo
                LIMIT 120""",
            tuple(params),
        )
        data = [
            {
                "periodo": r["periodo"],
                "precio_promedio": float(r["precio_promedio"] or 0),
                "precio_minimo": float(r["precio_minimo"] or 0),
                "precio_maximo": float(r["precio_maximo"] or 0),
                "registros": int(r["registros"]),
            }
            for r in (rows or [])
            if r.get("periodo")
        ]
        return {"success": True, "insumo": insumo, "data": data}
    except mysql.connector.Error:
        log.exception("Database error in historico_precios")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al consultar el histórico de precios.")


@router.get("/dashboard", tags=["APUs"])
async def get_dashboard_stats_endpoint() -> dict:
    try:
        return get_dashboard_stats()
    except mysql.connector.Error:
        log.exception("Database error in get_dashboard_stats")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al cargar las métricas del dashboard.")


@router.get("/projects", tags=["APUs"])
async def get_projects() -> dict:
    try:
        return {"projects": get_unique_projects()}
    except mysql.connector.Error:
        log.exception("Database error in get_projects")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al cargar la lista de proyectos.")


@router.get("/proyectos-mapus", tags=["APUs"])
async def listar_proyectos_mapus() -> dict:
    try:
        rows = execute_query(
            """SELECT p.id, p.id_proy, p.descripcion, p.presupuesto_total,
                      COALESCE(p.aiu_administracion, 15.00) AS aiu_administracion,
                      COALESCE(p.aiu_imprevistos, 3.00) AS aiu_imprevistos,
                      COALESCE(p.aiu_utilidad, 5.00) AS aiu_utilidad,
                      COALESCE(p.aiu_iva_utilidad, 19.00) AS aiu_iva_utilidad,
                      COUNT(ip.id) AS items_apu_cargados,
                      COUNT(CASE WHEN ip.aprobado_interventoria = 1 THEN 1 END) AS items_apu_aprobados,
                      COALESCE(SUM(ip.valor_presupuestado), 0) AS total_apu_cargado
               FROM proyectos p
               LEFT JOIN item_proyecto ip ON ip.proyecto = p.id AND ip.apu_solicitud_id IS NOT NULL
               GROUP BY p.id
               ORDER BY p.id DESC"""
        )
        return {"success": True, "proyectos": rows or []}
    except Exception:
        log.exception("Error listando proyectos MAPUS")
        raise HTTPException(status_code=500, detail="Error al cargar proyectos.")


@router.get("/proyectos-mapus/{proyecto_id}/items", tags=["APUs"])
async def detalle_items_proyecto(proyecto_id: int) -> dict:
    """Árbol de ítems del presupuesto de un proyecto. Los ítems cargados desde un APU
    aprobado (apu_solicitud_id no nulo) traen la fecha y el responsable de la firma legal."""
    from src.infrastructure.database.connection import execute_query

    try:
        rows = await asyncio.to_thread(
            execute_query,
            """SELECT ip.id, ip.parent_id, ip.nivel, ip.codigo, ip.nombre,
                      ip.unidad_medida, ip.cantidad_presupuestada, ip.valor_unitario,
                      ip.valor_presupuestado, ip.orden, ip.tipo_item, ip.apu_solicitud_id,
                      h.aprobado_por, h.aprobado_rol, h.aprobado_en
               FROM item_proyecto ip
               LEFT JOIN (
                   SELECT solicitud_id,
                          MAX(created_at) AS aprobado_en,
                          SUBSTRING_INDEX(GROUP_CONCAT(responsable_nombre ORDER BY created_at DESC), ',', 1) AS aprobado_por,
                          SUBSTRING_INDEX(GROUP_CONCAT(responsable_rol ORDER BY created_at DESC), ',', 1) AS aprobado_rol
                   FROM historial_aprobaciones
                   WHERE accion = 'aprobado_legal'
                   GROUP BY solicitud_id
               ) h ON h.solicitud_id = ip.apu_solicitud_id
               WHERE ip.proyecto = %s
               ORDER BY ip.orden, ip.id""",
            (proyecto_id,),
        )
        proy = await asyncio.to_thread(
            execute_query,
            """SELECT id, id_proy, descripcion, presupuesto_total,
                      COALESCE(aiu_administracion, 15.00) AS aiu_administracion,
                      COALESCE(aiu_imprevistos, 3.00) AS aiu_imprevistos,
                      COALESCE(aiu_utilidad, 5.00) AS aiu_utilidad,
                      COALESCE(aiu_iva_utilidad, 19.00) AS aiu_iva_utilidad
               FROM proyectos WHERE id = %s""",
            (proyecto_id,),
        )
        return {"success": True, "proyecto": (proy or [None])[0], "items": rows or []}
    except Exception:
        log.exception("Error obteniendo detalle del proyecto %d", proyecto_id)
        raise HTTPException(status_code=500, detail="Error al cargar el detalle del proyecto.")


class CrearProyectoRequest(BaseModel):
    id_proy: int
    descripcion: Optional[str] = ""
    presupuesto_total: Optional[float] = 0
    id_folder: str = "local"
    id_folder_bim: Optional[str] = None
    pdo_current_version_id: Optional[int] = None
    pdo_drive_subfolder_id: Optional[str] = None
    aiu_administracion: Optional[float] = 15.00
    aiu_imprevistos: Optional[float] = 3.00
    aiu_utilidad: Optional[float] = 5.00
    aiu_iva_utilidad: Optional[float] = 19.00


class ActualizarProyectoRequest(BaseModel):
    descripcion: Optional[str] = None
    presupuesto_total: Optional[float] = None
    aiu_administracion: Optional[float] = None
    aiu_imprevistos: Optional[float] = None
    aiu_utilidad: Optional[float] = None
    aiu_iva_utilidad: Optional[float] = None


@router.post("/proyectos-mapus", tags=["APUs"])
async def crear_proyecto(payload: CrearProyectoRequest) -> dict:
    try:
        execute_query(
            """INSERT INTO proyectos (id_proy, descripcion, presupuesto_total, id_folder, id_folder_bim,
                                      pdo_current_version_id, pdo_drive_subfolder_id,
                                      aiu_administracion, aiu_imprevistos, aiu_utilidad, aiu_iva_utilidad)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (payload.id_proy, payload.descripcion, payload.presupuesto_total, payload.id_folder,
             payload.id_folder_bim, payload.pdo_current_version_id, payload.pdo_drive_subfolder_id,
             payload.aiu_administracion, payload.aiu_imprevistos, payload.aiu_utilidad, payload.aiu_iva_utilidad),
            fetch=False,
        )
        project_id = execute_query("SELECT LAST_INSERT_ID() AS id")[0]["id"]
        log.info("Proyecto %d creado con AIU paramétrico: %s", project_id, payload.descripcion)
        return {"success": True, "id": project_id}
    except Exception:
        log.exception("Error creando proyecto")
        raise HTTPException(status_code=500, detail="Error al crear proyecto.")


@router.patch("/proyectos-mapus/{proyecto_id}", tags=["APUs"])
async def actualizar_proyecto(proyecto_id: int, payload: ActualizarProyectoRequest) -> dict:
    try:
        updates = []
        params = []
        if payload.descripcion is not None:
            updates.append("descripcion = %s")
            params.append(payload.descripcion)
        if payload.presupuesto_total is not None:
            updates.append("presupuesto_total = %s")
            params.append(payload.presupuesto_total)
        if payload.aiu_administracion is not None:
            updates.append("aiu_administracion = %s")
            params.append(payload.aiu_administracion)
        if payload.aiu_imprevistos is not None:
            updates.append("aiu_imprevistos = %s")
            params.append(payload.aiu_imprevistos)
        if payload.aiu_utilidad is not None:
            updates.append("aiu_utilidad = %s")
            params.append(payload.aiu_utilidad)
        if payload.aiu_iva_utilidad is not None:
            updates.append("aiu_iva_utilidad = %s")
            params.append(payload.aiu_iva_utilidad)

        if not updates:
            return {"success": True, "mensaje": "Sin cambios"}

        params.append(proyecto_id)
        execute_query(
            f"UPDATE proyectos SET {', '.join(updates)} WHERE id = %s",
            tuple(params),
            fetch=False,
        )
        return {"success": True, "id": proyecto_id}
    except Exception:
        log.exception("Error actualizando proyecto %d", proyecto_id)
        raise HTTPException(status_code=500, detail="Error al actualizar proyecto.")


@router.get("/proyectos-mapus/{proyecto_id}/apus", tags=["APUs"])
async def apus_del_proyecto(
    proyecto_id: int,
    limit: int = Query(200, ge=1, le=1000),
    offset: int = Query(0, ge=0),
) -> dict:
    """Filas del banco de APUs asignadas directamente a este proyecto (proyecto_id),
    complementa el árbol de item_proyecto que devuelve /proyectos-mapus/{id}/items."""
    try:
        rows = await asyncio.to_thread(
            execute_query,
            """SELECT id, item, items_descripcion, item_unidad, ciudad, contratista,
                      numero_contrato, precio_unitario, tipo_insumo, insumo_descripcion,
                      rendimiento_insumo, precio_unitario_apu, link_documento, created_at
               FROM apus
               WHERE proyecto_id = %s
               ORDER BY item, id
               LIMIT %s OFFSET %s""",
            (proyecto_id, limit, offset),
        )
        total = await asyncio.to_thread(
            execute_query,
            "SELECT COUNT(*) AS total FROM apus WHERE proyecto_id = %s",
            (proyecto_id,),
        )
        return {"success": True, "apus": rows or [], "total": (total or [{"total": 0}])[0]["total"]}
    except Exception:
        log.exception("Error obteniendo APUs del proyecto %d", proyecto_id)
        raise HTTPException(status_code=500, detail="Error al cargar los APUs del proyecto.")


class AsignarProyectoRequest(BaseModel):
    proyecto_id: int
    apu_ids: list[int]


@router.post("/asignar-proyecto", tags=["APUs"])
async def asignar_proyecto(
    payload: AsignarProyectoRequest,
    user: dict = Depends(require_role("analista")),
) -> dict:
    """Asigna filas del banco de APUs (por id) a un proyecto — usado al subir/actualizar
    APUs al banco para vincularlos a un proyecto ya creado."""
    if not payload.apu_ids:
        raise HTTPException(status_code=400, detail="Debe indicar al menos un APU.")
    try:
        proyecto = await asyncio.to_thread(
            execute_query, "SELECT id FROM proyectos WHERE id = %s", (payload.proyecto_id,)
        )
        if not proyecto:
            raise HTTPException(status_code=404, detail="Proyecto no encontrado.")

        marcadores = ",".join(["%s"] * len(payload.apu_ids))
        await asyncio.to_thread(
            execute_query,
            f"UPDATE apus SET proyecto_id = %s WHERE id IN ({marcadores})",
            (payload.proyecto_id, *payload.apu_ids),
            fetch=False,
        )
        log.info(
            "Proyecto %d asignado a %d APU(s) por %s (%s)",
            payload.proyecto_id, len(payload.apu_ids), user["nombre"], user["rol"],
        )
        return {"success": True, "asignados": len(payload.apu_ids)}
    except HTTPException:
        raise
    except Exception:
        log.exception("Error asignando proyecto %d a APUs", payload.proyecto_id)
        raise HTTPException(status_code=500, detail="Error al asignar el proyecto.")


@router.delete("/projects", tags=["APUs"])
async def delete_projects(nombre_proyecto: str = Query(..., min_length=1), user: dict = Depends(require_role("admin"))) -> dict:
    try:
        log.warning("Project deletion by %s (%s): %s", user["nombre"], user["rol"], nombre_proyecto)
        return await asyncio.to_thread(delete_project, nombre_proyecto)
    except mysql.connector.Error:
        log.exception("Database error in delete_projects")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al eliminar el proyecto.")
