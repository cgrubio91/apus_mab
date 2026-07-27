"""
Script: Cargar presupuesto de obra desde Excel a la DB.
Lee el archivo ACTA_30, crea el proyecto "Vías del Renacimiento"
y carga los capítulos (nivel 1) e ítems (nivel 2) en item_proyecto.
"""

import re
import logging
from decimal import Decimal

import openpyxl
from src.infrastructure.database.connection import get_db_connection

log = logging.getLogger("cargar_presupuesto")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

RUTA = r"C:\Users\cgrub\Downloads\invias 2026_1\Vias_del_Renacimiento_-_Acta_ACTA_30_generada_el_2026-07-27.xlsx"
HOJA = "ACTA_30_SegMAB"
NOMBRE_PROYECTO = "Vías del Renacimiento"


def normalizar(val):
    if val is None:
        return None
    s = str(val).strip().rstrip("_")
    return s if s else None


def es_capitulo(col_a):
    return col_a and re.match(r"^Capitulo\s+\d", str(col_a).strip(), re.IGNORECASE)


def es_resumen(col_a, col_e):
    a = (col_a or "").strip().upper()
    e = (col_e or "").strip().upper()
    keywords = ["SUBTOTAL", "REAJUSTE", "PLAN DE MANEJO", "PAGA", "AJUSTE A DISEÑOS", "TOTAL DE OBRA"]
    return any(k in a or k in e for k in keywords)


def main():
    wb = openpyxl.load_workbook(RUTA, data_only=True)
    ws = wb[HOJA]

    conn = get_db_connection()
    try:
        with conn.cursor(dictionary=True) as cur:
            # 1. Obtener siguiente id_proy
            cur.execute("SELECT COALESCE(MAX(id_proy), 0) + 1 AS next_id FROM proyectos")
            next_id_proy = cur.fetchone()["next_id"]

            # 2. Crear proyecto
            cur.execute(
                "INSERT INTO proyectos (id_proy, descripcion, id_folder) VALUES (%s, %s, %s)",
                (next_id_proy, NOMBRE_PROYECTO, "vias_renacimiento"),
            )
            proyecto_id = cur.lastrowid
            log.info("Proyecto creado: id=%d, id_proy=%d, desc=%s", proyecto_id, next_id_proy, NOMBRE_PROYECTO)

            # 3. Parsear Excel
            chapter_id = None
            orden = 0

            for r in range(4, ws.max_row + 1):
                col_a = normalizar(ws.cell(r, 1).value)
                col_b = normalizar(ws.cell(r, 2).value)
                col_c = normalizar(ws.cell(r, 3).value)
                col_d = normalizar(ws.cell(r, 4).value)
                col_e = normalizar(ws.cell(r, 5).value)
                col_f = ws.cell(r, 6).value
                col_g = ws.cell(r, 7).value
                col_h = normalizar(ws.cell(r, 8).value)
                col_i = ws.cell(r, 9).value
                col_j = ws.cell(r, 10).value
                col_k = ws.cell(r, 11).value

                # Saltar filas vacías
                if not col_a and not col_e:
                    continue

                # Saltar filas de resumen (SUBTOTAL, TOTAL, etc.)
                if es_resumen(col_a, col_e):
                    log.info("  [resumen] %s %s", col_a or "", col_e or "")
                    continue

                # Procesar capítulo
                if es_capitulo(col_a):
                    total = Decimal(str(col_k)) if col_k else Decimal(0)
                    cur.execute(
                        """INSERT INTO item_proyecto
                           (proyecto, nivel, codigo, nombre, valor_presupuestado, tipo_item, orden, aprobado_interventoria)
                           VALUES (%s, 1, %s, %s, %s, 'PREVISTO', %s, 1)""",
                        (proyecto_id, col_a, col_a, total, orden),
                    )
                    chapter_id = cur.lastrowid
                    orden += 1
                    log.info("  Capítulo %d: %s", chapter_id, col_a)
                    continue

                # Si no hay capítulo activo, saltar
                if chapter_id is None:
                    continue

                # Procesar ítem
                cantidad = Decimal(str(col_i)) if col_i else Decimal(0)
                vr_unitario = Decimal(str(col_j)) if col_j else Decimal(0)
                vr_total = Decimal(str(col_k)) if col_k else Decimal(0)
                ai = Decimal(str(col_f)) if col_f else None
                aiu = Decimal(str(col_g)) if col_g else None

                cur.execute(
                    """INSERT INTO item_proyecto
                       (proyecto, parent_id, nivel, codigo, especif_gral_raw, especif_part_raw,
                        grupo_ajuste_raw, nombre, unidad_medida, cantidad_presupuestada,
                        valor_unitario, valor_presupuestado, ai, aiu, tipo_item, orden, aprobado_interventoria)
                       VALUES (%s, %s, 2, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'PREVISTO', %s, 1)""",
                    (proyecto_id, chapter_id, col_a, col_b, col_c, col_d,
                     col_e or col_a, col_h, cantidad, vr_unitario, vr_total,
                     ai, aiu, orden),
                )
                orden += 1

            # 4. Actualizar presupuesto total del proyecto
            cur.execute(
                "SELECT SUM(valor_presupuestado) AS total FROM item_proyecto WHERE proyecto = %s AND nivel = 1",
                (proyecto_id,),
            )
            total_presupuesto = cur.fetchone()["total"] or Decimal(0)
            cur.execute(
                "UPDATE proyectos SET presupuesto_total = %s WHERE id = %s",
                (total_presupuesto, proyecto_id),
            )

            conn.commit()
            log.info("Carga completada: proyecto_id=%d, capítulos y %d ítems insertados, presupuesto_total=%s",
                     proyecto_id, orden, total_presupuesto)

    except Exception:
        conn.rollback()
        log.exception("Error durante la carga")
        raise
    finally:
        conn.close()
    wb.close()


if __name__ == "__main__":
    main()
